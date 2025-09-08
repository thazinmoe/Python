#!/usr/bin/env python3
import json
import os
import re
import sys
import warnings
from typing import Any, Dict, List, Optional, Iterable
from datetime import date, datetime, time

# Optional: silence noisy extension warnings from openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries, column_index_from_string
from openpyxl.styles.colors import COLOR_INDEX


# ---------- helpers ----------
def _apply_tint(hex_rgb: str, tint: Optional[float]) -> str:
    """Apply Excel tint ([-1,1]) to a '#RRGGBB' color.

    Positive values lighten; negative values darken.
    """
    if not hex_rgb:
        return hex_rgb
    if tint is None or tint == 0:
        return hex_rgb
    s = hex_rgb.lstrip('#')
    if len(s) != 6:
        return hex_rgb
    r = int(s[0:2], 16)
    g = int(s[2:4], 16)
    b = int(s[4:6], 16)

    def adj(ch: int) -> int:
        if tint < 0:
            v = ch * (1.0 + float(tint))
        else:
            v = ch * (1.0 - float(tint)) + 255.0 * float(tint)
        return max(0, min(255, int(round(v))))

    r, g, b = adj(r), adj(g), adj(b)
    return f"#{r:02X}{g:02X}{b:02X}"


def _build_theme_rgb_map(wb) -> Dict[int, str]:
    """Extract theme color mapping index->'#RRGGBB' from workbook theme.

    Tries both Workbook._theme (parsed) and Workbook.loaded_theme (raw XML).
    """
    mapping: Dict[int, str] = {}

    # Preferred: parsed theme object
    try:
        th = getattr(wb, "_theme", None)
        if th and getattr(th, "themeElements", None):
            cs = th.themeElements.clrScheme
            order = [
                "lt1", "dk1", "lt2", "dk2",
                "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
                "hlink", "folHlink",
            ]
            for i, key in enumerate(order):
                v = getattr(cs, key, None)
                if not v:
                    continue
                val = None
                if getattr(v, "srgbClr", None) is not None:
                    val = getattr(v.srgbClr, "val", None)
                elif getattr(v, "sysClr", None) is not None:
                    val = getattr(v.sysClr, "lastClr", None)
                if val and len(val) == 6:
                    mapping[i] = "#" + val.upper()
            if mapping:
                return mapping
    except Exception:
        pass

    # Fallback: parse theme XML
    try:
        theme_xml = getattr(wb, "loaded_theme", None)
        if theme_xml:
            from xml.etree import ElementTree as ET
            if isinstance(theme_xml, bytes):
                root = ET.fromstring(theme_xml)
            else:
                root = ET.fromstring(str(theme_xml))
            ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
            cs = root.find('.//a:themeElements/a:clrScheme', ns)
            if cs is not None:
                order = [
                    "lt1", "dk1", "lt2", "dk2",
                    "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
                    "hlink", "folHlink",
                ]
                for i, key in enumerate(order):
                    el = cs.find(f'a:{key}', ns)
                    if el is None:
                        continue
                    srgb = el.find('a:srgbClr', ns)
                    sysc = el.find('a:sysClr', ns)
                    val = None
                    if srgb is not None and 'val' in srgb.attrib:
                        val = srgb.attrib['val']
                    elif sysc is not None and 'lastClr' in sysc.attrib:
                        val = sysc.attrib['lastClr']
                    if val and len(val) == 6:
                        mapping[i] = "#" + val.upper()
    except Exception:
        pass

    return mapping


def color_to_hex(c, theme_map: Optional[Dict[int, str]] = None) -> Optional[str]:
    """Convert openpyxl Color or ARGB/HEX string to '#RRGGBB', resolving theme/tint.

    Handles:
    - Color.rgb (8/6 length)
    - Color.indexed via COLOR_INDEX mapping
    - Color.theme + tint via workbook theme map
    - Plain strings 'FF112233' / '112233'
    """
    if not c:
        return None

    # If it's an openpyxl Color-like object
    try:
        # Direct RGB
        rgb = getattr(c, "rgb", None)
        if rgb:
            s = str(rgb).upper()
            if len(s) == 8:
                return "#" + s[2:]
            if len(s) == 6:
                return "#" + s

        # Indexed palette fallback
        idx = getattr(c, "indexed", None)
        if idx is not None:
            try:
                i = int(idx)
                if 0 <= i < len(COLOR_INDEX):
                    pal = (COLOR_INDEX[i] or "").upper()
                    if len(pal) == 8:
                        return "#" + pal[2:]
                    if len(pal) == 6:
                        return "#" + pal
            except Exception:
                pass

        # Theme + optional tint
        theme = getattr(c, "theme", None)
        if theme is not None and theme_map:
            try:
                i = int(theme)
                base = theme_map.get(i)
                if base:
                    tint = getattr(c, "tint", None)
                    try:
                        tint = float(tint) if tint is not None else None
                    except Exception:
                        tint = None
                    return _apply_tint(base, tint)
            except Exception:
                pass
    except Exception:
        pass

    # Maybe it is a plain string like 'FFRRGGBB' or 'RRGGBB'
    try:
        s = str(c).upper()
        if len(s) == 8:
            return "#" + s[2:]
        if len(s) == 6:
            return "#" + s
    except Exception:
        pass

    return None


def font_to_obj(f, to_hex) -> Dict[str, Any]:
    return {
        "name": getattr(f, "name", None),
        "size": getattr(f, "size", None),
        "bold": getattr(f, "bold", None),
        "italic": getattr(f, "italic", None),
        "underline": bool(getattr(f, "underline", None)),
        "strike": getattr(f, "strike", None),
        "color": to_hex(getattr(f, "color", None)),
    }


def fill_to_obj(fill, to_hex) -> Optional[Dict[str, Any]]:
    if not fill or not getattr(fill, "fill_type", None):
        return None
    fg = getattr(fill, "fgColor", None)
    bg = getattr(fill, "bgColor", None)
    fg_hex = to_hex(getattr(fg, "color", None) or fg)
    bg_hex = to_hex(getattr(bg, "color", None) or bg)
    return {
        "type": fill.fill_type,
        "fgColor": fg_hex,
        "bgColor": bg_hex,
    }


def edge_to_obj(edge, to_hex):
    if not edge or not getattr(edge, "style", None):
        return None
    return {"style": edge.style, "color": to_hex(getattr(edge, "color", None))}


def border_to_obj(b, to_hex) -> Dict[str, Any]:
    return {
        "left": edge_to_obj(getattr(b, "left", None), to_hex),
        "right": edge_to_obj(getattr(b, "right", None), to_hex),
        "top": edge_to_obj(getattr(b, "top", None), to_hex),
        "bottom": edge_to_obj(getattr(b, "bottom", None), to_hex),
        "diagonal": edge_to_obj(getattr(b, "diagonal", None), to_hex),
    }


def alignment_to_obj(a) -> Dict[str, Any]:
    """Return a minimal alignment dict with only necessary fields.

    The export intentionally omits less commonly needed fields like
    wrapText, indent, textRotation, shrinkToFit. Only include keys that
    have non-None values to keep the JSON compact.
    """
    out: Dict[str, Any] = {}
    horiz = getattr(a, "horizontal", None)
    vert = getattr(a, "vertical", None)
    if horiz is not None:
        out["horizontal"] = horiz
    if vert is not None:
        out["vertical"] = vert
    return out


def rich_text_runs(cell, to_hex) -> Optional[List[Dict[str, Any]]]:
    """
    Best-effort extraction of inline rich text (varies by openpyxl version).
    """
    v = getattr(cell, "_value", None)
    runs = []
    candidates = None
    for attr in ("rich", "_rich", "r"):
        if hasattr(v, attr):
            val = getattr(v, attr)
            if isinstance(val, list):
                candidates = val
                break
    if not candidates:
        return None
    for run in candidates:
        text = getattr(run, "text", None) or getattr(run, "t", None) or getattr(run, "value", None)
        font = getattr(run, "font", None) or getattr(run, "rPr", None)
        runs.append({
            "text": text,
            "font": font_to_obj(font, to_hex) if font else None,
        })
    return runs or None


def extract_defined_names(wb) -> List[Dict[str, Any]]:
    """
    Return [{"name": str, "attr_text": str|None}] across openpyxl versions.
    Handles dict-like (3.x) and older list-like APIs.
    """
    out = []
    dn = getattr(wb, "defined_names", None)
    if dn is None:
        return out

    # Newer dict-like (preferred path)
    try:
        for k in dn:
            items = dn[k]
            if isinstance(items, list):
                for it in items:
                    out.append({"name": getattr(it, "name", k), "attr_text": getattr(it, "attr_text", None)})
            else:
                it = items
                out.append({"name": getattr(it, "name", k), "attr_text": getattr(it, "attr_text", None)})
        if out:
            return out
    except Exception:
        pass

    # Older list-like with .definedName
    try:
        for it in dn.definedName:
            out.append({"name": getattr(it, "name", None), "attr_text": getattr(it, "attr_text", None)})
        return out
    except Exception:
        pass

    # Fallback: iterate values
    try:
        vals = getattr(dn, "values", None)
        if callable(vals):
            vals = dn.values()
        if vals is not None:
            for it in vals:
                if isinstance(it, list):
                    for x in it:
                        out.append({"name": getattr(x, "name", None), "attr_text": getattr(x, "attr_text", None)})
                else:
                    out.append({"name": getattr(it, "name", None), "attr_text": getattr(it, "attr_text", None)})
    except Exception:
        pass

    return out


# ---------- core ----------
def _extract_sheet_info(ws, to_hex, ws_values=None) -> Dict[str, Any]:
    """Extract detailed info for a single worksheet into a serializable dict."""
    # sheet meta
    try:
        selection = ws.sheet_view.selection[0] if ws.sheet_view and ws.sheet_view.selection else None
        top_left = getattr(selection, "activeCell", None) if selection else None
        pane = getattr(ws.sheet_view, "pane", None)
        pane_dict = pane.__dict__ if pane else None
    except Exception:
        top_left, pane_dict = None, None

    sheet_meta = {
        "state": getattr(ws, "sheet_state", None),  # 'visible' | 'hidden' | 'veryHidden'
        "views": {
            "frozenPane": {
                "topLeftCell": top_left,
                "pane": pane_dict
            }
        }
    }

    # columns info
    columns = {}
    for col_dim in ws.column_dimensions.values():
        idx = getattr(col_dim, "index", None)
        if idx is None:
            continue
        # openpyxl stores ColumnDimension.index as a letter like 'A'
        try:
            col_idx = column_index_from_string(idx) if isinstance(idx, str) else int(idx)
        except Exception:
            # fallback best-effort
            col_idx = int(getattr(col_dim, "min", 0) or 0) or None
        letter = idx if isinstance(idx, str) else get_column_letter(col_idx) if col_idx else None
        if col_idx:
            columns[str(col_idx)] = {
                "letter": letter,
                "width": getattr(col_dim, "width", None),
                "hidden": getattr(col_dim, "hidden", None),
            }

    # rows info
    rows = {}
    for idx, row_dim in ws.row_dimensions.items():
        rows[str(idx)] = {
            "height": getattr(row_dim, "height", None),
            "hidden": getattr(row_dim, "hidden", None),
            "outlineLevel": getattr(row_dim, "outlineLevel", None),
        }

    # data validations
    validations = []
    dv_container = getattr(ws, "data_validations", None)
    if dv_container:
        for dv in getattr(dv_container, "dataValidation", []):
            validations.append({
                "type": getattr(dv, "type", None),
                "operator": getattr(dv, "operator", None),
                "allowBlank": getattr(dv, "allow_blank", None),
                "showErrorMessage": getattr(dv, "showErrorMessage", None),
                "errorTitle": getattr(dv, "errorTitle", None),
                "error": getattr(dv, "error", None),
                "formula1": getattr(dv, "formula1", None),
                "formula2": getattr(dv, "formula2", None),
                "sqref": str(getattr(dv, "sqref", "")),
            })

    # conditional formatting (basic metadata)
    cond_formats = []
    cf = getattr(ws, "conditional_formatting", None)
    try:
        cf_rules = getattr(cf, "cf_rules", {}) if cf else {}
        for ref, rules in cf_rules.items():
            for rule in rules:
                cond_formats.append({
                    "ref": ref,
                    "type": getattr(rule, "type", None),
                    "operator": getattr(rule, "operator", None),
                    "formula": getattr(rule, "formula", None),
                    "dxf": True if getattr(rule, "dxf", None) else False,
                })
    except Exception:
        # IF openpyxl stripped extensions, this will be empty anyway
        pass

    sheet_info: Dict[str, Any] = {
        "meta": sheet_meta,
        "dimension": ws.calculate_dimension(),  # e.g. 'A1:F42'
        "merged": [str(rng) for rng in ws.merged_cells.ranges],
        "tables": [],
        "columns": columns,
        "rows": rows,
        "validations": validations,
        "conditionalFormats": cond_formats,
        "cells": {},
        "hyperlinks": {},
    }

    # tables (ListObjects)
    tables = getattr(ws, "tables", None) or getattr(ws, "_tables", {})
    if isinstance(tables, dict):
        for t in tables.values():
            sheet_info["tables"].append({
                "name": getattr(t, "displayName", getattr(t, "name", None)),
                "ref": getattr(t, "ref", None),
            })
    elif isinstance(tables, list):
        for t in tables:
            sheet_info["tables"].append({
                "name": getattr(t, "displayName", getattr(t, "name", None)),
                "ref": getattr(t, "ref", None),
            })

    # iterate only the declared used range to avoid huge sparse loops
    try:
        dim = ws.calculate_dimension()
        min_col, min_row, max_col, max_row = range_boundaries(dim)
    except Exception:
        min_row, min_col, max_row, max_col = ws.min_row, ws.min_column, ws.max_row, ws.max_column
    for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for c in row:
            addr = c.coordinate
            # Always record hyperlink metadata (even for visually empty cells)
            if getattr(c, "hyperlink", None) is not None:
                sheet_info["hyperlinks"][addr] = {
                    "target": getattr(c.hyperlink, "target", None),
                    "display": c.value if isinstance(c.value, str) else None,
                    "tooltip": getattr(c.hyperlink, "tooltip", None),
                }

            # Decide whether to include this cell even if empty: keep if it has
            # either a font color or a fill (background) color.
            v = c.value
            # If we have a companion data_only sheet, fetch cached/calculated value
            v_calc = None
            try:
                if ws_values is not None:
                    v_calc = ws_values[addr].value
            except Exception:
                v_calc = None
            is_empty = (v is None) or (isinstance(v, str) and v == "")
            has_color_or_bg = False
            if is_empty:
                try:
                    # Check for background (fill) color
                    bg_col = None
                    if getattr(c, "fill", None) and getattr(c.fill, "fill_type", None):
                        fg = getattr(c.fill, "fgColor", None)
                        bg = getattr(c.fill, "bgColor", None)
                        fg_hex = to_hex(getattr(fg, "color", None) or fg)
                        bg_hex = to_hex(getattr(bg, "color", None) or bg)
                        bg_col = fg_hex or bg_hex

                    # Only consider background fills to avoid default font color noise
                    has_color_or_bg = bool(bg_col)
                except Exception:
                    has_color_or_bg = False

            # Skip if empty AND no color/background styling
            if is_empty and not has_color_or_bg:
                continue

            # Prefer the cached/calculated value for formula cells.
            # Preserve the formula string separately.
            cell_obj: Dict[str, Any] = {}
            try:
                is_formula = isinstance(v, str) and v.startswith("=")
            except Exception:
                is_formula = False

            if is_formula:
                cell_obj["formula"] = v
                cell_obj["value"] = v_calc
            else:
                cell_obj["value"] = v

            if getattr(c, "number_format", None):
                cell_obj["numFmt"] = c.number_format
            if getattr(c, "alignment", None):
                _al = alignment_to_obj(c.alignment)
                if _al:
                    cell_obj["alignment"] = _al
            if getattr(c, "font", None):
                cell_obj["font"] = font_to_obj(c.font, to_hex)
            if getattr(c, "fill", None) and getattr(c.fill, "fill_type", None):
                cell_obj["fill"] = fill_to_obj(c.fill, to_hex)
            if getattr(c, "border", None):
                cell_obj["border"] = border_to_obj(c.border, to_hex)

            # rich text (best effort)
            rt = rich_text_runs(c, to_hex)
            if rt:
                cell_obj["richText"] = rt

            # Convenience: expose commonly requested color fields at top-level
            # - color: font color (text color)
            # - backgroundColor: solid fill color when available (fgColor preferred)
            try:
                font_color = None
                if getattr(c, "font", None) and getattr(c.font, "color", None) is not None:
                    font_color = to_hex(c.font.color)
                cell_obj["color"] = font_color
            except Exception:
                cell_obj["color"] = None

            try:
                bg_color = None
                if getattr(c, "fill", None) and getattr(c.fill, "fill_type", None):
                    # Prefer fgColor for solid fills; fall back to bgColor
                    fg = getattr(c.fill, "fgColor", None)
                    bg = getattr(c.fill, "bgColor", None)
                    fg_hex = to_hex(getattr(fg, "color", None) or fg)
                    bg_hex = to_hex(getattr(bg, "color", None) or bg)
                    bg_color = fg_hex or bg_hex
                cell_obj["backgroundColor"] = bg_color
            except Exception:
                cell_obj["backgroundColor"] = None

            sheet_info["cells"][addr] = cell_obj

    return sheet_info


def dump_workbook(path: str, only_sheets: Optional[Iterable[str]] = None) -> Dict[str, Any]:
    """Dump workbook to a dict. If only_sheets is provided, include only those sheets."""
    # Load twice: one for full formatting + formulas, one with data_only for cached values
    wb = load_workbook(path, data_only=False, read_only=False)
    wb_vals = load_workbook(path, data_only=True, read_only=False)
    theme_map = _build_theme_rgb_map(wb)

    # local resolver to capture theme_map
    def to_hex(c):
        return color_to_hex(c, theme_map)
    out: Dict[str, Any] = {"file": path, "sheets": {}}

    # workbook-level defined names
    out["definedNames"] = extract_defined_names(wb)

    names_filter = set(only_sheets) if only_sheets else None
    for ws in wb.worksheets:
        if names_filter and ws.title not in names_filter:
            continue
        # Match the corresponding sheet from the data_only workbook if available
        ws_vals = None
        try:
            ws_vals = wb_vals[ws.title]
        except Exception:
            ws_vals = None
        sheet_info = _extract_sheet_info(ws, to_hex, ws_values=ws_vals)
        out["sheets"][ws.title] = sheet_info

    return out


def _safe_filename(name: str) -> str:
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", name.strip())
    return s or "sheet"


def main():
    # Minimal argument parsing without external dependencies
    import argparse

    parser = argparse.ArgumentParser(
        description="Extract rich Excel formatting and data to JSON",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("input", help="Path to input .xlsx file")
    parser.add_argument("output", help="Output .json file or directory when using --split-sheets")
    parser.add_argument(
        "--sheet",
        dest="sheet",
        metavar="NAME",
        help="Export only this sheet into the output JSON (single file).",
    )
    parser.add_argument(
        "--split-sheets",
        dest="split_sheets",
        action="store_true",
        help="Write one JSON per sheet into the output directory",
    )

    args = parser.parse_args()

    # Helper for JSON serialization
    def _default(o):
        if isinstance(o, (datetime, date, time)):
            return o.isoformat()
        return str(o)

    inp = args.input
    outp = args.output

    if args.split_sheets:
        # Output path must be a directory
        out_dir = outp
        os.makedirs(out_dir, exist_ok=True)

        # Load workbook once and iterate sheets
        wb = load_workbook(inp, data_only=False, read_only=False)
        wb_vals = load_workbook(inp, data_only=True, read_only=False)
        theme_map = _build_theme_rgb_map(wb)

        def to_hex(c):
            return color_to_hex(c, theme_map)

        defined_names = extract_defined_names(wb)

        # Optional filter
        names_filter = {args.sheet} if args.sheet else None
        count = 0
        for ws in wb.worksheets:
            if names_filter and ws.title not in names_filter:
                continue
            data = {
                "file": inp,
                "definedNames": defined_names,
                "sheets": {ws.title: _extract_sheet_info(ws, to_hex, ws_values=wb_vals[ws.title] if ws.title in wb_vals.sheetnames else None)},
            }
            filename = _safe_filename(f"{ws.title}.json")
            dst = os.path.join(out_dir, filename)
            with open(dst, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=_default)
            count += 1
            print(f"Wrote {dst}")

        if count == 0:
            if args.sheet:
                print(f"Error: sheet '{args.sheet}' not found in workbook.")
            else:
                print("No sheets found to export.")
            sys.exit(2)
        return

    # Single output file path
    only = [args.sheet] if args.sheet else None
    data = dump_workbook(inp, only_sheets=only)
    with open(outp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=_default)
    print(f"Wrote {outp}")


if __name__ == "__main__":
    main()
